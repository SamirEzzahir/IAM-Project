"""Excel exports for assignment results and the reusable PCO catalog."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def _finish(workbook: Workbook, widths: list[int]) -> bytes:
    sheet = workbook.active
    fill = PatternFill("solid", fgColor="343A40")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def assignment_results_excel(job: dict) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Résultats Affectation"
    sheet.append([
        "#", "Login", "SPL", "PCO", "brin", "Motif", "Durée (s)",
        "Port MSAN", "Message",
    ])
    for index, row in enumerate(job.get("results", []), 1):
        sheet.append([
            row.get("excel_row") or index,
            row.get("login") or job.get("login") or "",
            row.get("spl") or job.get("spl") or "",
            row.get("pco") or "",
            row.get("selected_port") or "",
            row.get("status_label") or "",
            row.get("duration_seconds"),
            row.get("msan_port") or "",
            row.get("message") or "",
        ])
    return _finish(workbook, [8, 20, 25, 30, 10, 18, 14, 38, 65])


def available_pcos_excel(job: dict) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PCO disponibles"
    sheet.append(["SPL", "PCO", "brin", "État", "Date du contrôle"])
    for row in job.get("available_pcos", []):
        sheet.append([
            job.get("spl") or "",
            row.get("pco") or "",
            row.get("brin") or "",
            row.get("status_label") or "",
            row.get("checked_at") or "",
        ])
    return _finish(workbook, [25, 32, 12, 18, 28])


def renseigner_results_excel(job: dict) -> bytes:
    workbook = Workbook(); sheet = workbook.active; sheet.title = "Renseigner PCOs"
    sheet.append(["#", "Entrée", "Login", "Source Login", "Type recherche constitution", "SPL/SRO", "PCO", "brin", "Port (Nom Usuel + NE)", "État", "Durée (s)", "Message"])
    for row in job.get("results", []):
        sheet.append([row.get("excel_row"), row.get("input"), row.get("login"), row.get("source"), row.get("constitution_search_mode"), row.get("constitution_spl"), row.get("constitution_pco"), row.get("constitution_brin"), row.get("msan_port"), row.get("status_label"), row.get("duration_seconds"), row.get("message")])
    return _finish(workbook, [8, 22, 22, 22, 25, 28, 30, 12, 35, 18, 14, 60])


def commandes_results_excel(job: dict) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Informations terrain"
    sheet.append([
        "CMD", "Nom splitter", "Port PCO", "Nom PCO", "Modèle ONT",
        "Client contacté", "Distance branchement (m)", "État",
        "Durée (s)", "Message",
    ])
    for row in job.get("results", []):
        sheet.append([
            row.get("cmd"), row.get("nom_splitter"), row.get("port_pco"),
            row.get("nom_pco"), row.get("modele_ont"),
            row.get("client_contacte"), row.get("distance_branchement"),
            row.get("status_label"), row.get("duration_seconds"),
            row.get("message"),
        ])
    return _finish(workbook, [18, 18, 14, 24, 20, 20, 28, 18, 14, 60])
