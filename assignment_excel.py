"""Excel/CSV inputs for batch assignment and MSAN port-to-SPL mapping."""

from __future__ import annotations

import csv
import io
import json
import re
from io import BytesIO
from pathlib import Path
from threading import Lock

from openpyxl import load_workbook

from bulk_excel import cell_text, header_key
from pco_logic import parse_spl

MAX_ASSIGNMENT_ROWS = 2000
MAX_MAPPING_ROWS = 10000
MSAN_MAPPING_LOCK = Lock()


def _excel_rows(content: bytes):
    try:
        sheet = load_workbook(BytesIO(content), data_only=True, read_only=True).active
        return list(sheet.iter_rows(values_only=True))
    except Exception as exc:
        raise ValueError("Le fichier Excel est invalide ou illisible.") from exc


def parse_assignment_workbook(content: bytes) -> list[dict]:
    rows = _excel_rows(content)
    if not rows:
        raise ValueError("Le fichier Excel est vide.")
    keys = [header_key(value) for value in rows[0]]
    try:
        login_index = next(i for i, key in enumerate(keys) if key in {"LOGIN", "LOGINCLIENT"})
        spl_index = next(i for i, key in enumerate(keys) if key in {"SPL", "SPLITTER", "SRO", "SPLITTEROUSRO"})
    except StopIteration as exc:
        raise ValueError("Colonnes obligatoires : Login et SPL.") from exc

    result = []
    for excel_row, values in enumerate(rows[1:], 2):
        login = cell_text(values[login_index] if login_index < len(values) else "")
        spl = cell_text(values[spl_index] if spl_index < len(values) else "").upper()
        if not login and not spl:
            continue
        error = None
        if not login:
            error = "Login absent."
        else:
            try:
                parse_spl(spl)
            except ValueError as exc:
                error = str(exc)
        result.append({"excel_row": excel_row, "login": login, "spl": spl, "validation_error": error})
        if len(result) > MAX_ASSIGNMENT_ROWS:
            raise ValueError(f"Le fichier dépasse {MAX_ASSIGNMENT_ROWS} lignes.")
    if not result:
        raise ValueError("Aucune ligne Login/SPL trouvée.")
    return result


def parse_msan_mapping(content: bytes, suffix: str) -> list[dict]:
    if suffix.lower() == ".csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ValueError("Le fichier CSV doit être encodé en UTF-8.") from exc
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=";,\t,")
            source = list(csv.reader(io.StringIO(text), dialect))
        except csv.Error as exc:
            raise ValueError("Le format du fichier CSV est invalide.") from exc
    else:
        source = _excel_rows(content)
    if not source:
        raise ValueError("Le fichier de correspondance est vide.")
    keys = [header_key(value) for value in source[0]]
    try:
        port_index = next(i for i, key in enumerate(keys) if key in {"CARTE", "PORT", "PORTMSAN", "CARTEPORT"})
        spl_index = next(i for i, key in enumerate(keys) if key in {"SPL", "SPLITTER", "SRO", "SPLITTEROUSRO"})
    except StopIteration as exc:
        raise ValueError("Colonnes obligatoires : Carte et Splitter ou SRO.") from exc
    mappings = []
    for values in source[1:]:
        port = cell_text(values[port_index] if port_index < len(values) else "")
        spl = cell_text(values[spl_index] if spl_index < len(values) else "").upper()
        if port and spl:
            mappings.append({"port": port, "spl": spl})
            if len(mappings) > MAX_MAPPING_ROWS:
                raise ValueError(f"Le fichier dépasse {MAX_MAPPING_ROWS} correspondances.")
    if not mappings:
        raise ValueError("Aucune correspondance Carte/SPL trouvée.")
    return mappings


def save_msan_mapping(path: Path, mappings: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(".tmp")
    with MSAN_MAPPING_LOCK:
        temporary.write_text(
            json.dumps(mappings, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        temporary.replace(path)


def resolve_msan_spl(path: Path, port: str) -> str | None:
    if not path.exists():
        return None
    wanted = re.sub(r"\s+", "", str(port)).upper()
    with MSAN_MAPPING_LOCK:
        mappings = json.loads(path.read_text(encoding="utf-8"))
    for item in mappings:
        if re.sub(r"\s+", "", item.get("port", "")).upper() == wanted:
            return item.get("spl")
    return None
