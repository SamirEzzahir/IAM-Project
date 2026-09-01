import unittest
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from bulk_excel import (
    build_full_pco,
    cell_text,
    choose_full_pco,
    derive_pco_location,
    effective_brin,
    locate_bulk_columns,
    parse_bulk_workbook,
    write_bulk_results,
)


class BulkExcelTests(unittest.TestCase):
    def test_locates_confirmed_excel_columns_with_duplicate_pco_headers(self):
        headers = [
            "Commande   GPON\u00a0",
            "ONT",
            "Login\u00a0",
            "ODF",
            "PCO",
            "PCO",
            "brin",
            "Pose PCO(O/N)",
        ]
        columns = locate_bulk_columns(headers)
        self.assertEqual(columns["command"], 0)
        self.assertEqual(columns["login"], 2)
        self.assertEqual(columns["odf"], 3)
        self.assertEqual(columns["pco_indices"], [4, 5])
        self.assertEqual(columns["brin"], 6)

    def test_prefers_full_pco_over_short_duplicate_column(self):
        self.assertEqual(
            choose_full_pco(["OFOF-ZO-7122/2", "7122/2"]),
            "OFOF-ZO-7122/2",
        )

    def test_derives_exact_odf_and_zr_from_full_pco(self):
        self.assertEqual(
            derive_pco_location("OFOF-ZO-7122/2"),
            ("OFOF", "OFOF-ZO"),
        )
        self.assertEqual(
            derive_pco_location("OFOF92-ZO-1422/2"),
            ("OFOF92", "OFOF92-ZO"),
        )

    def test_excel_numeric_command_keeps_no_decimal_suffix(self):
        self.assertEqual(cell_text(101479263.0), "101479263")

    def test_builds_full_pco_from_odf_zr_and_short_pco(self):
        self.assertEqual(
            build_full_pco("OFOF-ZO", "2711"),
            ("OFOF", "OFOF-ZO", "OFOF-ZO-2711"),
        )

    def test_split_pco_maps_excel_brins_five_to_eight_onto_one_to_four(self):
        self.assertEqual(effective_brin("OFOF-ZO-2711/1", "5"), "1")
        self.assertEqual(effective_brin("OFOF-ZO-2711/2", "8"), "4")
        self.assertEqual(effective_brin("OFOF-ZO-2711", "8"), "8")

    def test_parses_real_workbook_using_exact_four_confirmed_values(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "Commande GPON", "Login", "ODF", "PCO", "brin", "ETAT",
        ])
        sheet.append([
            "DFOIWC00207739", "I10268094", "OFOF-ZO", "7122/2", 8, "",
        ])
        stream = BytesIO()
        workbook.save(stream)

        rows = parse_bulk_workbook(stream.getvalue())

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["command"], "DFOIWC00207739")
        self.assertEqual(rows[0]["login"], "I10268094")
        self.assertEqual(rows[0]["pco"], "OFOF-ZO-7122/2")
        self.assertEqual(rows[0]["brin"], "8")
        self.assertEqual(rows[0]["target_brin"], "4")
        self.assertEqual(rows[0]["odf"], "OFOF")
        self.assertEqual(rows[0]["zr"], "OFOF-ZO")

    def test_result_export_keeps_previous_login_and_final_status(self):
        rows = [{
            "excel_row": 2,
            "command": "DFOIWC00207739",
            "login": "I10268094",
            "pco": "OFOF-ZO-7122/2",
            "brin": "8",
            "target_brin": "4",
        }]
        results = [{
            "search_mode": "CMD",
            "previous_login": "OLDLOGIN",
            "spl": "OFOF-ZO-113.16",
            "status_label": "Muté",
            "port_spl": "OFOF-ZO-113.16",
            "msan_port": "MHOu-Fe-MourabitineERAC1--2C2:0-0-18-8",
            "message": "Mutation terminée.",
            "checked_at": "2026-08-28T12:00:00+00:00",
        }]
        with TemporaryDirectory() as directory:
            path = Path(directory) / "result.xlsx"
            write_bulk_results(path, rows, results)
            workbook = load_workbook(path, data_only=True)
            sheet = workbook["Résultats Mutation"]
            self.assertEqual(sheet["F2"].value, "4")
            self.assertEqual(sheet["G2"].value, "CMD")
            self.assertEqual(sheet["H2"].value, "OLDLOGIN")
            self.assertEqual(sheet["I2"].value, "OFOF-ZO-113.16")
            self.assertEqual(sheet["J2"].value, "Muté")
            self.assertEqual(sheet["K2"].value, "OFOF-ZO-113.16")
            self.assertEqual(sheet["L2"].value, "MHOu-Fe-MourabitineERAC1--2C2:0-0-18-8")


if __name__ == "__main__":
    unittest.main()
