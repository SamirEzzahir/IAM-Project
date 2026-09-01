import csv
import io
import unittest
from io import BytesIO
from threading import Event

import app
from openpyxl import load_workbook


class AppExportTests(unittest.TestCase):
    def test_sidebar_uses_real_links_as_navigation_fallback(self):
        response = app.app.test_client().get("/")
        html = response.get_data(as_text=True)
        self.assertIn('href="#tab-config" data-tab="config"', html)
        self.assertIn('href="#tab-available" data-tab="available"', html)
        self.assertIn("app.js?v=20260901-2", html)
        self.assertIn('name="assignmentMode" value="single"', html)
        self.assertIn('id="assignDownloadBtn"', html)
        self.assertIn('id="downloadAvailableExcel"', html)

    def test_javascript_has_executable_mime_type_with_nosniff(self):
        response = app.app.test_client().get("/static/app.js")
        try:
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.mimetype, "application/javascript")
            self.assertEqual(response.headers["X-Content-Type-Options"], "nosniff")
        finally:
            response.close()

    def test_available_csv_writes_one_line_per_free_brin(self):
        job_id = "csv-test"
        app.jobs[job_id] = {
            "job_id": job_id, "kind": "CHECK", "spl": "OFOF-ZO-113.1",
            "odf": "OFOF", "zr": "OFOF-ZO", "status": "COMPLETED",
            "total": 3, "completed_count": 3, "results": [{
                "status": "AVAILABLE", "pco": "OFOF-ZO-1111",
                "pco_exists": True, "free_ports": ["2", "4"],
                "checked_at": "2026-01-01",
            }, {
                "status": "SKIPPED", "pco": "OFOF-ZO-1111/1",
                "pco_exists": None, "free_ports": [],
            }, {
                "status": "SKIPPED", "pco": "OFOF-ZO-1111/2",
                "pco_exists": None, "free_ports": [],
            }], "logs": [], "run_event": Event(), "stop_event": Event(),
            "thread": None, "output_path": None,
        }
        try:
            response = app.app.test_client().get(f"/api/check/{job_id}/available.csv")
            rows = list(csv.reader(io.StringIO(response.data.decode("utf-8-sig")), delimiter=";"))
            self.assertEqual(rows[1][3:6], ["OFOF-ZO-1111", "2", "Disponible"])
            self.assertEqual(rows[2][3:6], ["OFOF-ZO-1111", "4", "Disponible"])

            excel_response = app.app.test_client().get(
                f"/api/check/{job_id}/available.xlsx"
            )
            self.assertEqual(excel_response.status_code, 200)
            workbook = load_workbook(BytesIO(excel_response.data), read_only=True)
            try:
                excel_rows = list(workbook.active.iter_rows(values_only=True))
                self.assertEqual(
                    excel_rows[0],
                    ("SPL", "PCO", "brin", "État", "Date du contrôle"),
                )
                self.assertEqual(
                    excel_rows[1],
                    ("OFOF-ZO-113.1", "OFOF-ZO-1111", "2", "Disponible", "2026-01-01"),
                )
            finally:
                workbook.close()
                excel_response.close()
        finally:
            app.jobs.pop(job_id, None)

    def test_assignment_excel_uses_normalized_columns(self):
        job_id = "assignment-export-test"
        app.jobs[job_id] = {
            "job_id": job_id, "kind": "BATCH_ASSIGNMENT",
            "status": "COMPLETED", "total": 1, "completed_count": 1,
            "created_at": "2026-09-01", "updated_at": "2026-09-01",
            "results": [{
                "excel_row": 1, "login": "logintest", "spl": "ODF-ZO-111.1",
                "pco": "ODF-ZO-111/1", "selected_port": "3",
                "status": "ASSIGNED", "status_label": "Affecté",
                "duration_seconds": 30.2,
                "msan_port": "GHI-FF-AinKadous:0-0-3-0", "message": "ok",
            }],
            "logs": [], "run_event": Event(), "stop_event": Event(),
            "thread": None, "output_path": None,
        }
        try:
            response = app.app.test_client().get(
                f"/api/assign/{job_id}/result.xlsx"
            )
            self.assertEqual(response.status_code, 200)
            workbook = load_workbook(BytesIO(response.data), read_only=True)
            try:
                rows = list(workbook.active.iter_rows(values_only=True))
                self.assertEqual(
                    rows[0],
                    ("#", "Login", "SPL", "PCO", "brin", "Motif", "Durée (s)", "Port MSAN", "Message"),
                )
                self.assertEqual(rows[1][1:9], (
                    "logintest", "ODF-ZO-111.1", "ODF-ZO-111/1", "3",
                    "Affecté", 30.2, "GHI-FF-AinKadous:0-0-3-0", "ok",
                ))
            finally:
                workbook.close()
                response.close()
        finally:
            app.jobs.pop(job_id, None)


if __name__ == "__main__":
    unittest.main()
