"""Local CMD-to-Login lookup for Degroupage FTTH workbooks."""

from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from bulk_excel import cell_text, header_key


def parse_degroupage_workbook(content: bytes) -> dict[str, dict[str, str]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        sheet = workbook["LoginDegroupageFTTH"] if "LoginDegroupageFTTH" in workbook.sheetnames else workbook.active
    except Exception as exc:
        raise ValueError("Le fichier Degroupage est invalide ou illisible.") from exc
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Le fichier Degroupage est vide.")
    headers = [header_key(value) for value in rows[0]]
    try:
        command_index = next(i for i, key in enumerate(headers) if key in {"CMD", "COMMANDE", "COMMANDEGPON"})
        login_index = next(i for i, key in enumerate(headers) if key in {"LOGIN", "LOGINCLIENT"})
    except StopIteration as exc:
        raise ValueError("Colonnes obligatoires : CMD et Login.") from exc
    nd_index = next((i for i, key in enumerate(headers) if key == "ND"), None)
    lookup: dict[str, dict[str, str]] = {}
    for values in rows[1:]:
        command = cell_text(values[command_index] if command_index < len(values) else "").upper()
        if not command or command in lookup:
            continue
        lookup[command] = {
            "login": cell_text(values[login_index] if login_index < len(values) else ""),
            "nd": cell_text(values[nd_index] if nd_index is not None and nd_index < len(values) else ""),
        }
    if not lookup:
        raise ValueError("Aucune ligne CMD/Login trouvée dans le fichier Degroupage.")
    return lookup


def save_degroupage_lookup(path: Path, lookup: dict[str, dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(".tmp")
    temporary.write_text(json.dumps(lookup, ensure_ascii=False), encoding="utf-8")
    temporary.replace(path)


def load_degroupage_lookup(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return value if isinstance(value, dict) else {}
