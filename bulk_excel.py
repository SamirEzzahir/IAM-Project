"""Excel parsing and result export for Bulk Mutation CMD&Login."""

from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path

from wimtech_parser import normalize


MAX_BULK_ROWS = 2000


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).replace("\u00a0", " ").split()).strip()


def header_key(value) -> str:
    return re.sub(r"[^A-Z0-9]+", "", normalize(cell_text(value)))


def locate_bulk_columns(headers: list) -> dict:
    keys = [header_key(value) for value in headers]

    def first_index(accepted: set[str], label: str) -> int:
        for index, key in enumerate(keys):
            if key in accepted:
                return index
        raise ValueError(f"Colonne Excel obligatoire introuvable : {label}.")

    command_index = first_index(
        {"COMMANDEGPON", "COMMANDE", "CMD", "NCOMMANDE"},
        "Commande GPON",
    )
    login_index = first_index({"LOGIN", "LOGINCLIENT"}, "Login")
    odf_index = first_index({"ODF", "ZR", "SRO", "ZRSRO"}, "ODF")
    brin_index = first_index({"BRIN", "PORT", "PORTFIBRE"}, "brin")
    pco_indices = [index for index, key in enumerate(keys) if key == "PCO"]
    if not pco_indices:
        raise ValueError("Colonne Excel obligatoire introuvable : PCO.")

    return {
        "command": command_index,
        "login": login_index,
        "odf": odf_index,
        "pco_indices": pco_indices,
        "brin": brin_index,
    }


def choose_full_pco(values: list[str]) -> str:
    candidates = [cell_text(value).upper() for value in values if cell_text(value)]
    if not candidates:
        return ""
    full = [value for value in candidates if "-" in value]
    return max(full or candidates, key=len)


def derive_pco_location(pco: str) -> tuple[str, str]:
    """Derive ODF and ZR from an exact full PCO supplied by Excel."""

    value = cell_text(pco).upper().rstrip("-")
    if "-" not in value:
        raise ValueError("Le PCO doit être complet, par exemple OFOF-ZO-7122/2.")
    zr, suffix = value.rsplit("-", 1)
    if not suffix or not re.search(r"\d", suffix):
        raise ValueError(f"Format PCO invalide : {value}.")

    if "-ZO" in zr:
        odf = zr.split("-ZO", 1)[0]
    elif zr.endswith("ZO") and len(zr) > 2:
        odf = zr[:-2].rstrip("-")
    else:
        odf = zr.split("-", 1)[0]
    if not odf or not zr:
        raise ValueError(f"Impossible de déterminer ODF/ZR depuis {value}.")
    return odf, zr


def build_full_pco(odf_value: str, pco_value: str) -> tuple[str, str, str]:
    """Build the exact WimTech ODF, ZR/SRO and PCO from Excel values."""

    location = cell_text(odf_value).upper().rstrip("-")
    pco = cell_text(pco_value).upper().strip().strip("-")
    if not location:
        raise ValueError("ODF absent.")
    if not pco:
        raise ValueError("PCO absent.")

    # A complete PCO is still accepted, but the ODF/ZR column remains the
    # authoritative location for the new five-column import format.
    full_pco = pco if "-" in pco else f"{location}-{pco}"
    if "-ZO" in location:
        odf = location.split("-ZO", 1)[0]
        zr = location
    elif location.endswith("ZO") and len(location) > 2:
        odf = location[:-2].rstrip("-")
        zr = location
    else:
        odf, zr = derive_pco_location(full_pco)
    if not odf or not zr:
        raise ValueError(f"Format ODF/ZR invalide : {location}.")
    return odf, zr, full_pco


def effective_brin(pco: str, brin: str) -> str:
    """Map Excel brins 5-8 to ports 1-4 for split /1 and /2 PCOs."""

    value = int(cell_text(brin))
    if value < 1 or value > 8:
        raise ValueError("brin invalide : la valeur doit être comprise entre 1 et 8.")
    if re.search(r"/(?:1|2)$", cell_text(pco)) and value >= 5:
        value -= 4
    return str(value)


def parse_bulk_workbook(content: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl n’est pas installé. Relancez run.bat.") from exc

    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError("Le fichier Excel est invalide ou illisible.") from exc

    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = list(next(iterator))
    except StopIteration as exc:
        raise ValueError("Le fichier Excel est vide.") from exc
    columns = locate_bulk_columns(headers)

    rows: list[dict] = []
    for excel_row, values in enumerate(iterator, start=2):
        values = list(values)
        command = cell_text(values[columns["command"]] if columns["command"] < len(values) else "")
        login = cell_text(values[columns["login"]] if columns["login"] < len(values) else "")
        odf_input = cell_text(values[columns["odf"]] if columns["odf"] < len(values) else "")
        brin = cell_text(values[columns["brin"]] if columns["brin"] < len(values) else "")
        pco = choose_full_pco([
            values[index] if index < len(values) else ""
            for index in columns["pco_indices"]
        ])
        if not any((command, login, odf_input, pco, brin)):
            continue

        error = None
        odf = ""
        zr = ""
        target_brin = ""
        if not command and not login:
            error = "Commande GPON et Login absents."
        elif not pco:
            error = "PCO absent."
        elif not odf_input:
            error = "ODF absent."
        elif not brin or not re.fullmatch(r"\d+", brin):
            error = "brin absent ou invalide."
        else:
            try:
                odf, zr, pco = build_full_pco(odf_input, pco)
                target_brin = effective_brin(pco, brin)
            except ValueError as exc:
                error = str(exc)

        rows.append({
            "excel_row": excel_row,
            "command": command,
            "login": login,
            "odf_input": odf_input,
            "pco": pco,
            "brin": str(int(brin)) if brin.isdigit() else brin,
            "target_brin": target_brin,
            "odf": odf,
            "zr": zr,
            "validation_error": error,
        })
        if len(rows) > MAX_BULK_ROWS:
            raise ValueError(f"Le fichier dépasse la limite de {MAX_BULK_ROWS} lignes.")

    if not rows:
        raise ValueError("Aucune ligne exploitable trouvée dans le fichier Excel.")
    return rows


def write_bulk_results(path: Path, rows: list[dict], results: list[dict]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl n’est pas installé. Relancez run.bat.") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Résultats Mutation"
    headers = [
        "Ligne Excel", "Commande GPON", "Login demandé", "PCO", "brin",
        "Brin utilisé", "Recherche utilisée", "Login précédent", "SPL",
        "État", "Message", "Date",
    ]
    sheet.append(headers)
    for row, result in zip(rows, results):
        sheet.append([
            row.get("excel_row"), row.get("command"), row.get("login"),
            row.get("pco"), row.get("brin"), row.get("target_brin"),
            result.get("search_mode"), result.get("previous_login"),
            result.get("spl"), result.get("status_label"),
            result.get("message"), result.get("checked_at"),
        ])

    fill = PatternFill("solid", fgColor="343A40")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [12, 20, 20, 28, 10, 12, 20, 22, 24, 18, 65, 25]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
